"""通用执行报告 → Excel。

适配「统一报告契约」(apps.api.main._build_executive_summary 的输出),所有工具通用:
  概览 / 问题清单 / 风险 / 阻碍 / 测试用例 / 截图索引 (+ 按工具的特化表)

输入:
  summary : _build_executive_summary(report, tool) 的结果(已归一)
  report  : 原始报告 dict(取 strengths / meta.screenshots / cases 原始字段)
  tool    : TOOL_CATALOG 条目 {id,name,icon}
  meta    : {run_id, produced_at, model, project_name, project_code}
  cfg     : 按工具的差异 {title, issue_sheet, issue_kind, biz, cases_primary}

沿用 seo_excel 同款视觉(配色 / 表头 / 状态三色)。
"""
from __future__ import annotations

import io
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from packages.reporting.seo_excel import (
    _F, _header, _row, _status_fill, _title_row, _widths,
    FONT_CELL, FONT_LBL, C_HEADER,
)

# 优先级配色
PRI_FILL = {
    "P0": PatternFill("solid", fgColor="FFC7CE"),
    "P1": PatternFill("solid", fgColor="FFEB9C"),
    "P2": PatternFill("solid", fgColor="FFF2CC"),
    "P3": PatternFill("solid", fgColor="EFEFEF"),
}
SEV_LABEL = {"critical": "致命", "high": "高", "medium": "中", "low": "低", "info": "信息"}
SEV_STATUS = {"critical": "不通过", "high": "警告", "medium": "警告", "low": "", "info": ""}
OWNER_LABEL = {"backend": "后端", "frontend": "前端", "product": "产品", "test": "测试",
               "devops": "运维", "security": "安全", "data": "数据", "design": "设计"}
_VERDICT_STATUS = {"pass": "通过", "warn": "警告", "fail": "不通过", "skip": "未产出"}


def _pri_status(pri: str) -> str:
    p = str(pri or "").upper()
    return "不通过" if p == "P0" else ("警告" if p == "P1" else "")


def _flat(v: Any) -> Any:
    """把 list/dict 拍平成可写入单元格的字符串(数字保留原类型)。"""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "是" if v else "否"
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, list):
        return "\n".join(str(_flat(x)) for x in v if x not in (None, ""))
    if isinstance(v, dict):
        return str(v.get("assert") or v.get("expected") or v.get("action")
                   or v.get("text") or v.get("title") or v)
    return str(v)


def _row_safe(ws, row: int, values: list[Any], status_cols=None, zebra=False):
    """_row 的安全版:写入前把每个值拍平(防 list/dict 触发 openpyxl 报错)。"""
    _row(ws, row, [_flat(v) for v in values], status_cols=status_cols, zebra=zebra)


def _ws(wb, name: str):
    """创建 sheet,净化非法字符(/ \\ ? * [ ] :)并截断到 31 字符(Excel 限制)。"""
    safe = re.sub(r'[/\\?*\[\]:]', '·', str(name)).strip()[:31] or "Sheet"
    return wb.create_sheet(safe)


def _steps_text(steps: Any) -> str:
    if isinstance(steps, list):
        out = []
        for i, s in enumerate(steps, 1):
            if isinstance(s, dict):
                txt = s.get("action") or s.get("step") or str(s)
                out.append(f"{s.get('order', i)}、{txt}")
            else:
                s = str(s).strip()
                out.append(s if (s and s[0].isdigit()) else f"{i}、{s}")
        return "\n".join(out)
    return str(steps or "")


def _expected_text(exp: Any) -> str:
    if isinstance(exp, list):
        out = []
        for e in exp:
            if isinstance(e, dict):
                out.append(str(e.get("assert") or e.get("expected") or e))
            else:
                out.append(str(e))
        return "\n".join(out)
    return str(exp or "")


def _autoheight(ws, row: int, text: str, per_line_chars: int = 40, base: int = 22, cap: int = 200):
    n = 0
    for seg in str(text or "").split("\n"):
        n += max(1, len(seg) // per_line_chars + 1)
    ws.row_dimensions[row].height = max(base, min(cap, n * 15 + 7))


# ══════════════════════════════════════════════════════════════════
def build_exec_xlsx(summary: dict[str, Any], report: dict[str, Any],
                    tool: dict[str, Any], meta: dict[str, Any],
                    cfg: dict[str, Any] | None = None) -> bytes:
    cfg = cfg or {}
    wb = Workbook()
    wb.remove(wb.active)

    n = [1]  # sheet 序号计数器(闭包可变)

    def _idx() -> str:
        s = f"{n[0]:02d}"
        n[0] += 1
        return s

    raw_cases = report.get("cases") if isinstance(report.get("cases"), list) else []
    if not raw_cases:
        for sub in (report.get("substeps") or {}).values():
            if isinstance(sub, dict) and sub.get("cases"):
                raw_cases = sub["cases"]
                break

    _sheet_overview(wb, _idx(), summary, report, tool, meta, cfg)

    # step2:用例表作为主表,排在问题清单之前
    if cfg.get("cases_primary") and raw_cases:
        _sheet_cases(wb, f"{_idx()} 测试用例", raw_cases)

    _sheet_issues(wb, f"{_idx()} {cfg.get('issue_sheet', '问题清单')}", summary, cfg)

    if summary.get("风险结论"):
        _sheet_risks(wb, f"{_idx()} 风险清单", summary["风险结论"])
    if summary.get("阻碍"):
        _sheet_blockers(wb, f"{_idx()} 阻碍·待澄清", summary["阻碍"], cfg)
    if raw_cases and not cfg.get("cases_primary"):
        _sheet_cases(wb, f"{_idx()} 测试用例", raw_cases)

    shots = [s for s in ((report.get("meta") or {}).get("screenshots") or [])
             if not s.get("error") and s.get("filename")]
    if shots:
        if cfg.get("h5_matrix"):
            _sheet_h5_matrix(wb, f"{_idx()} 浏览器·视口矩阵", shots)
        _sheet_shots(wb, f"{_idx()} 截图索引", shots)

    if not wb.sheetnames:
        _ws(wb, "概览")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sheet_overview(wb, idx, summary, report, tool, meta, cfg):
    ws = _ws(wb, f"{idx} 概览")
    _widths(ws, [16, 26, 22, 20, 22])
    tname = tool.get("name") or tool.get("id") or "报告"
    title = cfg.get("title") or f"{tname}报告"
    concl = summary.get("测试结论", {}) or {}
    sev = summary.get("严重度分布", {}) or {}
    cases = summary.get("用例执行", {}) or {}
    pri = cases.get("按优先级", {}) or {}
    sub = (f"工具:{tname}  |  运行:{meta.get('run_id', '')[:8]}  |  "
           f"时间:{meta.get('produced_at', '') or '—'}  |  模型:{meta.get('model', '') or '—'}")
    _title_row(ws, title, 5, sub)

    strengths = report.get("strengths") or []
    if isinstance(strengths, str):
        strengths = [strengths]
    top_fix = "；".join(
        str(i.get("fix") or "") for i in (summary.get("问题描述") or [])[:3] if i.get("fix"))

    rows = [
        ("测试结论", concl.get("判定") or "—", _VERDICT_STATUS.get(concl.get("level"), "")),
        ("一句话结论", concl.get("summary") or "(见各表明细)", ""),
        ("问题统计", f"致命 {sev.get('critical', 0)} · 高 {sev.get('high', 0)} · "
                     f"中 {sev.get('medium', 0)} · 低 {sev.get('low', 0)} · 信息 {sev.get('info', 0)}", ""),
        ("用例统计", f"共 {cases.get('总数', 0)} 条　(P0 {pri.get('P0', 0)} / "
                     f"P1 {pri.get('P1', 0)} / P2 {pri.get('P2', 0)} / P3 {pri.get('P3', 0)})", ""),
        ("亮点 / 已验证", "；".join(str(s) for s in strengths) if strengths else "(无)", ""),
        ("后续动作", top_fix or "(见问题清单)", ""),
    ]
    r = 4
    for lbl, val, status in rows:
        lc = ws.cell(r, 1, lbl)
        lc.font = FONT_LBL
        lc.fill = PatternFill("solid", fgColor="E8F1F2")
        lc.alignment = Alignment(vertical="top", wrap_text=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        vc = ws.cell(r, 2, str(val))
        vc.font = FONT_CELL
        vc.alignment = Alignment(wrap_text=True, vertical="top")
        if status:
            fill, font = _status_fill(status)
            if fill:
                vc.fill, vc.font = fill, font
        _autoheight(ws, r, val, per_line_chars=60, base=26)
        r += 1


def _sheet_issues(wb, title, summary, cfg):
    ws = _ws(wb, title)
    issues = summary.get("问题描述") or []
    biz = bool(cfg.get("biz"))
    if biz:
        headers = ["编号", "优先级", "需求点 / 模块", "问题 / 现象", "详情 / 现状", "影响", "期望 / 应满足", "待确认 · 建议"]
        widths = [12, 8, 20, 34, 38, 26, 30, 38]
    else:
        headers = ["编号", "优先级", "严重度", "模块 / 位置", "问题标题", "现状", "期望",
                   "修复建议", "复现步骤", "验收标准", "影响面", "负责人", "工时", "关联用例", "证据"]
        widths = [12, 8, 9, 22, 32, 36, 30, 38, 32, 26, 24, 9, 7, 18, 24]
    _widths(ws, widths)
    r = _title_row(ws, title.split(" ", 1)[-1] + "(按 严重度 × 优先级 排序)", len(headers))
    _header(ws, r, headers)
    ws.freeze_panes = ws.cell(r + 1, 1).coordinate
    r += 1
    if not issues:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
        ws.cell(r, 1, "(本次未识别到具体问题)").font = Font(name=_F, size=10, italic=True, color="6B7280")
        return
    pref = cfg.get("issue_prefix", "ISS")
    for i, it in enumerate(issues, 1):
        iid = it.get("issue_id") or f"{pref}-{i:02d}"
        pri = str(it.get("priority") or "P2").upper()
        if biz:
            vals = [iid, pri, it.get("module") or "", it.get("title") or "",
                    it.get("current") or "", it.get("impact") or "",
                    it.get("expected") or "", it.get("fix") or ""]
            status_cols = {2: _pri_status(pri)}
        else:
            sevk = it.get("severity") or "medium"
            repro = "\n".join(f"{j}. {s}" for j, s in enumerate(it.get("repro") or [], 1))
            owner = OWNER_LABEL.get(it.get("owner"), it.get("owner") or "")
            vals = [iid, pri, SEV_LABEL.get(sevk, sevk), it.get("module") or "",
                    it.get("title") or "", it.get("current") or "", it.get("expected") or "",
                    it.get("fix") or "", repro, it.get("accept") or "", it.get("impact") or "",
                    owner, it.get("hours") or "", "、".join(it.get("cases") or []), it.get("evidence") or ""]
            status_cols = {2: _pri_status(pri), 3: SEV_STATUS.get(sevk, "")}
        _row_safe(ws, r, vals, status_cols=status_cols)
        if not biz:
            pc = ws.cell(r, 2)
            if pri in PRI_FILL and not status_cols.get(2):
                pc.fill = PRI_FILL[pri]
        _autoheight(ws, r, (it.get("current") or "") + (it.get("fix") or ""), per_line_chars=70)
        r += 1


def _sheet_risks(wb, title, risks):
    ws = _ws(wb, title)
    _widths(ws, [12, 10, 40, 40, 40])
    r = _title_row(ws, "风险清单", 5)
    _header(ws, r, ["编号", "严重度", "风险", "影响", "原因 / 说明"])
    r += 1
    for i, rk in enumerate(risks, 1):
        sevk = rk.get("severity") or "medium"
        _row_safe(ws, r, [rk.get("id") or f"R-{i:02d}", SEV_LABEL.get(sevk, sevk),
                     rk.get("title") or "", rk.get("impact") or "", rk.get("why") or ""],
             status_cols={2: SEV_STATUS.get(sevk, "")})
        _autoheight(ws, r, rk.get("title", ""), per_line_chars=38)
        r += 1


def _sheet_blockers(wb, title, blockers, cfg):
    ws = _ws(wb, title)
    _widths(ws, [12, 36, 44, 44, 12, 8])
    label = "待澄清问题" if cfg.get("biz") else "阻碍项"
    r = _title_row(ws, f"{label}(需先解决才能继续)", 6)
    _header(ws, r, ["编号", "标题", "为何阻碍 / 待确认", "如何解开 / 建议", "负责方", "工时"])
    r += 1
    for i, b in enumerate(blockers, 1):
        owner = OWNER_LABEL.get(b.get("owner_role"), b.get("owner_role") or "")
        _row_safe(ws, r, [b.get("id") or f"B-{i:02d}", b.get("title") or "",
                     b.get("why_blocking") or "", b.get("what_to_unblock") or "",
                     owner, b.get("hours") or ""])
        _autoheight(ws, r, (b.get("why_blocking") or "") + (b.get("what_to_unblock") or ""), per_line_chars=42)
        r += 1


def _sheet_cases(wb, title, cases):
    ws = _ws(wb, title)
    headers = ["用例编号", "所属模块", "用例标题", "优先级", "用例类型", "前置条件",
               "测试步骤", "预期结果", "自动化", "执行结果", "实际结果"]
    _widths(ws, [14, 14, 30, 8, 10, 24, 40, 36, 10, 12, 22])
    r = _title_row(ws, f"测试用例(共 {len(cases)} 条 · 执行结果/实际结果留空给人填)", len(headers))
    _header(ws, r, headers)
    ws.freeze_panes = ws.cell(r + 1, 1).coordinate
    r += 1
    if not cases:
        ws.cell(r, 1, "(本报告无测试用例)").font = Font(name=_F, size=10, italic=True, color="6B7280")
        return
    for i, c in enumerate(cases, 1):
        if not isinstance(c, dict):
            continue
        pri = str(c.get("priority") or "P2").upper()
        vals = [c.get("id") or c.get("case_id") or f"TC-{i:03d}", c.get("module") or "",
                c.get("title") or c.get("name") or c.get("scenario") or "", pri,
                c.get("type") or "", c.get("preconditions") or "",
                _steps_text(c.get("steps")), _expected_text(c.get("expected") or c.get("expected_result")),
                c.get("automation_tag") or c.get("automation") or "", "", ""]
        _row_safe(ws, r, vals, status_cols={4: _pri_status(pri)})
        if pri in PRI_FILL and pri not in ("P0", "P1"):
            ws.cell(r, 4).fill = PRI_FILL[pri]
        _autoheight(ws, r, _steps_text(c.get("steps")), per_line_chars=40)
        r += 1


def _sheet_h5_matrix(wb, title, shots):
    """H5 适配:按 视口 × 页面 的实拍矩阵。"""
    ws = _ws(wb, title)
    _widths(ws, [30, 16, 14, 12, 12])
    r = _title_row(ws, "浏览器 / 视口适配实测矩阵", 5)
    _header(ws, r, ["页面 / URL", "视口", "尺寸", "问题数", "判定"])
    r += 1
    for s in shots:
        ic = s.get("issue_count", 0) or 0
        _row_safe(ws, r, [s.get("url") or "—", s.get("viewport") or "—",
                     f"{s.get('width', '?')}×{s.get('height', '?')}", ic,
                     "不通过" if ic else "通过"],
             status_cols={5: "不通过" if ic else "通过"})
        r += 1


def _sheet_shots(wb, title, shots):
    ws = _ws(wb, title)
    _widths(ws, [40, 16, 14, 10, 36])
    r = _title_row(ws, f"截图证据索引(共 {len(shots)} 张 · 图片见 evidence/screenshots 目录)", 5)
    _header(ws, r, ["页面 / URL", "视口", "尺寸", "问题数", "文件名"])
    r += 1
    for s in shots:
        _row_safe(ws, r, [s.get("url") or "—", s.get("viewport") or "—",
                     f"{s.get('width', '?')}×{s.get('height', '?')}", s.get("issue_count", 0) or 0,
                     s.get("annotated_filename") or s.get("filename") or ""])
        r += 1
