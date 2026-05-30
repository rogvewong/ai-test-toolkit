"""SEO 深度审计报告 → Excel(10 sheet),对标用户给定的模板结构与配色。

输入:
  seo_data : SeoAuditData.to_dict() —— 确定性采集的事实层
  synthesis: LLM 综合分析 {overview:{评定,核心问题,后续动作}, issues:[...], strengths:[...]}
  meta     : 运行元信息 {site, run_id, model, crawled_at, baseline?:上次run的seo_data}
输出: xlsx bytes
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── 配色 ──
C_HEADER = "1F4E5F"      # 表头深青
C_TITLE = "0E7C7B"       # 标题青
C_PASS_BG, C_PASS_FG = "C6EFCE", "006100"
C_WARN_BG, C_WARN_FG = "FFEB9C", "9C5700"
C_FAIL_BG, C_FAIL_FG = "FFC7CE", "9C0006"
C_ZEBRA = "F4F8F9"

_F = "微软雅黑"
FONT_TITLE = Font(name=_F, size=15, bold=True, color="FFFFFF")
FONT_SUB = Font(name=_F, size=10, color="FFFFFF")
FONT_HEAD = Font(name=_F, size=10, bold=True, color="FFFFFF")
FONT_CELL = Font(name=_F, size=10, color="1F2937")
FONT_LBL = Font(name=_F, size=10, bold=True, color="1F4E5F")
THIN = Side(style="thin", color="D9E2E8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _status_fill(val: str):
    v = str(val or "")
    if any(k in v for k in ("通过", "✅", "✓", "良好", "优秀")) and "不通过" not in v:
        return PatternFill("solid", fgColor=C_PASS_BG), Font(name=_F, size=10, color=C_PASS_FG, bold=True)
    if any(k in v for k in ("不通过", "❌", "差")):
        return PatternFill("solid", fgColor=C_FAIL_BG), Font(name=_F, size=10, color=C_FAIL_FG, bold=True)
    if any(k in v for k in ("警告", "⚠️", "◑", "需改进", "待")):
        return PatternFill("solid", fgColor=C_WARN_BG), Font(name=_F, size=10, color=C_WARN_FG, bold=True)
    return None, None


def _title_row(ws, text: str, ncols: int, sub: str | None = None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, text)
    c.font = FONT_TITLE
    c.fill = PatternFill("solid", fgColor=C_TITLE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    r = 2
    if sub:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        sc = ws.cell(2, 1, sub)
        sc.font = Font(name=_F, size=9, color="5B6B73")
        sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 20
        r = 3
    return r + 1  # 下一可写行(留一空行)


def _header(ws, row: int, headers: list[str]):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i, h)
        c.font = FONT_HEAD
        c.fill = PatternFill("solid", fgColor=C_HEADER)
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 22


def _row(ws, row: int, values: list[Any], status_cols: dict[int, Any] | None = None, zebra=False):
    for i, v in enumerate(values, 1):
        c = ws.cell(row, i, v if v is not None else "")
        c.font = FONT_CELL
        c.alignment = WRAP
        c.border = BORDER
        if zebra:
            c.fill = PatternFill("solid", fgColor=C_ZEBRA)
    if status_cols:
        for col, val in status_cols.items():
            fill, font = _status_fill(val)
            if fill:
                cc = ws.cell(row, col)
                cc.fill, cc.font = fill, font
                cc.alignment = CENTER


def _widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════════════════════════════
def build_seo_xlsx(seo_data: dict[str, Any], synthesis: dict[str, Any], meta: dict[str, Any]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    _sheet_overview(wb, seo_data, synthesis, meta)
    _sheet_tracking(wb, seo_data, meta)
    _sheet_tech(wb, seo_data)
    _sheet_cwv(wb, seo_data)
    _sheet_structured(wb, seo_data)
    _sheet_links(wb, seo_data)
    _sheet_templates(wb, seo_data)
    _sheet_issues(wb, synthesis)
    _sheet_pages(wb, seo_data)
    _sheet_criteria(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sheet_overview(wb, d, syn, meta):
    ws = wb.create_sheet("01 综合总览")
    _widths(ws, [16, 22, 22, 22, 22, 22])
    s = d.get("summary", {})
    site = meta.get("site") or d.get("base_url", "")
    sub = (f"测试日期:{meta.get('crawled_at','')}  |  站点:{site}  |  "
           f"爬取:{s.get('pages_crawled',0)} 页(200 OK {s.get('pages_ok',0)})  |  模型:{meta.get('model','')}")
    _title_row(ws, f"{meta.get('site_name','站点')} SEO 深度测试报告", 6, sub)
    ov = syn.get("overview", {}) if isinstance(syn, dict) else {}
    rows = [("综合评定", ov.get("评定") or ov.get("verdict_summary") or "(待 AI 综合)"),
            ("核心问题", ov.get("核心问题") or ov.get("core_issues") or ""),
            ("后续动作", ov.get("后续动作") or ov.get("next_actions") or "")]
    r = 4
    for lbl, val in rows:
        lc = ws.cell(r, 1, lbl); lc.font = FONT_LBL; lc.alignment = Alignment(vertical="top", wrap_text=True)
        lc.fill = PatternFill("solid", fgColor="E8F1F2"); lc.border = BORDER
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        vc = ws.cell(r, 2, str(val)); vc.font = FONT_CELL; vc.alignment = WRAP; vc.border = BORDER
        ws.row_dimensions[r].height = max(60, min(220, str(val).count("\n") * 16 + len(str(val)) // 6))
        r += 1


def _sheet_tracking(wb, d, meta):
    ws = wb.create_sheet("02 V2→V3 改进追踪")
    _widths(ws, [30, 22, 26, 12, 30])
    r = _title_row(ws, "改进追踪(与上次基线对比)", 5)
    base = meta.get("baseline")
    _header(ws, r, ["检查项", "上次(基线)", "本次", "变化", "备注"]); r += 1
    if not base:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        c = ws.cell(r, 1, "首次运行,暂无历史基线可对比。下次对同一站点再跑时,本表将自动列出每项指标的修复/回归/保持。")
        c.font = Font(name=_F, size=10, italic=True, color="9C5700")
        c.fill = PatternFill("solid", fgColor=C_WARN_BG); c.alignment = LEFT; c.border = BORDER
        ws.row_dimensions[r].height = 36
        return
    cur, prev = d.get("summary", {}), (base.get("summary", {}) or {})
    items = [
        ("爬取页数", prev.get("pages_crawled"), cur.get("pages_crawled")),
        ("死链", prev.get("dead_links"), cur.get("dead_links")),
        ("标题重复", prev.get("title_dup"), cur.get("title_dup")),
        ("描述<50字", prev.get("desc_short"), cur.get("desc_short")),
        ("图片Alt<80%", prev.get("alt_below80"), cur.get("alt_below80")),
        ("EN页title含中文", prev.get("en_title_cn"), cur.get("en_title_cn")),
        ("H1 i18n占位符", prev.get("h1_i18n_leak"), cur.get("h1_i18n_leak")),
        ("JSON-LD缺失", prev.get("jsonld_missing"), cur.get("jsonld_missing")),
    ]
    for name, pv, cv in items:
        if pv is None and cv is None:
            continue
        delta = "—"
        if isinstance(pv, (int, float)) and isinstance(cv, (int, float)):
            delta = "✅ 改进" if cv < pv else ("❌ 回归" if cv > pv else "✓ 保持")
        _row(ws, r, [name, pv, cv, delta, ""], status_cols={4: delta}); r += 1


def _sheet_tech(wb, d):
    ws = wb.create_sheet("03 站点级与技术SEO")
    _widths(ws, [26, 46, 12, 40])
    r = _title_row(ws, "站点级与技术 SEO", 4)
    _header(ws, r, ["检查项", "实测结果", "状态", "建议"]); r += 1
    t = d.get("tech", {}); sm = d.get("sitemap", {}); s = d.get("summary", {})
    def st(ok, warn=False): return "不通过" if not ok and not warn else ("警告" if warn else "通过")
    rows = [
        ("HTTPS / TLS", "全站 HTTPS" if t.get("https") else "非 HTTPS", st(t.get("https")), "保持" if t.get("https") else "全站改 HTTPS"),
        ("HTTP 版本", f"{t.get('http_version','?')}" + (f"; alt-svc {t.get('alt_svc','')[:30]}" if t.get("alt_svc") else ""), "通过", "保持"),
        ("压缩", t.get("content_encoding") or "无", st(bool(t.get("content_encoding"))), "保持" if t.get("content_encoding") else "开启 gzip/br"),
        ("robots.txt", f"HTTP {sm.get('robots',{}).get('status','?')}" + ("，含 AI 爬虫拦截" if sm.get('robots',{}).get('has_ai_blocker') else ""), st(sm.get("robots",{}).get("status")==200), "保持"),
        ("robots Sitemap 协议", sm.get("protocol_in_robots") or "未声明", "警告" if sm.get("protocol_in_robots")=="http" else "通过", "改 https://" if sm.get("protocol_in_robots")=="http" else "保持"),
        ("sitemap.xml", f"HTTP {sm.get('index_status','?')}，{sm.get('url_count',0)} URL", st(sm.get("index_status")==200), "保持" if sm.get("index_status")==200 else "新建标准 sitemap"),
        ("sitemap lastmod", f"{sm.get('with_lastmod',0)}/{sm.get('url_count',0)} 含 lastmod", st(sm.get("with_lastmod",0)>0, warn=(0<sm.get('with_lastmod',0)<sm.get('url_count',1))), "补全 lastmod"),
        ("sitemap 覆盖率", f"{sm.get('crawl_coverage_pct',0)}% 爬取页在 sitemap", st(sm.get("crawl_coverage_pct",0)>=80, warn=30<=sm.get("crawl_coverage_pct",0)<80), "扩展分片覆盖"),
        ("canonical 缺失", f"{sum(1 for p in d.get('pages',[]) if p.get('canonical_issue')=='缺')} 页缺 canonical", st(s.get("pages_ok",0)>0 and sum(1 for p in d.get('pages',[]) if p.get('canonical_issue')=='缺')==0, warn=True), "为所有页补自指 canonical"),
        ("HSTS", t.get("hsts") or "缺失", st(bool(t.get("hsts")), warn=not t.get("hsts")), "补 max-age=31536000; includeSubDomains"),
        ("CSP", t.get("csp") or "缺失", "警告" if not t.get("csp") else "通过", "按业务补 CSP"),
        ("X-Content-Type-Options", t.get("x_content_type_options") or "缺失", "警告" if not t.get("x_content_type_options") else "通过", "补 nosniff"),
        ("X-Frame-Options", t.get("x_frame_options") or "缺失", st(bool(t.get("x_frame_options")), warn=not t.get("x_frame_options")), "保持/补 SAMEORIGIN"),
        ("断链/孤儿页", f"{s.get('dead_links',0)} 死链，{s.get('orphan_pages',0)} 孤儿页", st(s.get("dead_links",0)==0 and s.get("orphan_pages",0)==0, warn=s.get("dead_links",0)>0), "保持" if s.get("dead_links",0)==0 else "修复死链"),
        ("内链锚文本", f"{s.get('generic_anchor_pct',0)}% 通用/空锚文本", st(s.get("generic_anchor_pct",100)<5, warn=s.get("generic_anchor_pct",100)<20), "保持" if s.get("generic_anchor_pct",100)<5 else "替换为实义锚文本"),
    ]
    for name, result, status, sug in rows:
        _row(ws, r, [name, result, status, sug], status_cols={3: status}); r += 1


def _sheet_cwv(wb, d):
    ws = wb.create_sheet("04 Core Web Vitals")
    _widths(ws, [16, 10, 10, 10, 10, 10, 10, 16, 10])
    r = _title_row(ws, "Core Web Vitals(真实浏览器渲染实测)", 9)
    _header(ws, r, ["模板", "TTFB", "FCP", "LCP", "CLS", "Load", "SSR比", "资源数/KB", "判定"]); r += 1
    cwv = d.get("cwv", {})
    if not cwv or cwv.get("error"):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        c = ws.cell(r, 1, "未采集到 Core Web Vitals" + (f"({cwv.get('error')})" if cwv.get("error") else ""))
        c.font = Font(name=_F, size=10, italic=True, color="9C5700"); c.alignment = LEFT; r += 1
    else:
        for tmpl, m in cwv.items():
            if not isinstance(m, dict):
                continue
            def ms(x): return f"{x}ms" if isinstance(x, (int, float)) else "—"
            _row(ws, r, [tmpl, ms(m.get("ttfb")), ms(m.get("fcp")), ms(m.get("lcp")),
                         m.get("cls") if m.get("cls") is not None else "—", ms(m.get("load")),
                         m.get("ssr_ratio") if m.get("ssr_ratio") is not None else "—",
                         f"{m.get('resources','?')} / {m.get('transferKB','?')}KB", m.get("verdict", "未采")],
                 status_cols={9: m.get("verdict", "")}); r += 1
    # 阈值参考
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    c = ws.cell(r, 1, "阈值参考"); c.font = FONT_LBL; c.fill = PatternFill("solid", fgColor="E8F1F2"); r += 1
    _header(ws, r, ["指标", "良好(通过)", "需改进(警告)", "差(不通过)", "", "", "", "", ""]); r += 1
    for name, g, w, b in [("LCP", "≤2500ms", "2500-4000ms", ">4000ms"), ("CLS", "≤0.1", "0.1-0.25", ">0.25"),
                          ("INP", "≤200ms", "200-500ms", ">500ms"), ("TTFB", "≤800ms", "800-1800ms", ">1800ms"),
                          ("FCP", "≤1800ms", "1800-3000ms", ">3000ms")]:
        _row(ws, r, [name, g, w, b]); r += 1


def _sheet_structured(wb, d):
    ws = wb.create_sheet("05 结构化数据")
    _widths(ws, [22, 12, 34, 36])
    dist = d.get("jsonld_dist", {})
    total = sum(dist.values())
    r = _title_row(ws, "结构化数据(JSON-LD)校验", 4)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ws.cell(r, 1, f"全站 JSON-LD 类型分布(共 {total} 项)"); c.font = FONT_LBL
    c.fill = PatternFill("solid", fgColor="E8F1F2"); r += 1
    _header(ws, r, ["@type", "数量", "覆盖页", "建议"]); r += 1
    USES = {"BreadcrumbList": "面包屑导航", "VideoObject": "视频详情", "WebPage": "通用页面",
            "Article": "文章", "Product": "商品", "WebSite": "站点信息", "Organization": "组织",
            "SearchAction": "站内搜索框", "ComicSeries": "漫画系列", "ImageObject": "图片"}
    if not dist:
        _row(ws, r, ["(全站未检出 JSON-LD)", 0, "0", "建议为详情页/文章补 schema.org 结构化数据"]); r += 1
    else:
        for t, cnt in list(dist.items())[:20]:
            _row(ws, r, [t, cnt, "—", USES.get(t, "继续保持")]); r += 1
    # 富媒体卡可达性(简版)
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ws.cell(r, 1, "富媒体卡机会(基于检出的 @type)"); c.font = FONT_LBL
    c.fill = PatternFill("solid", fgColor="E8F1F2"); r += 1
    _header(ws, r, ["@type", "卡片机会", "状态", "建议"]); r += 1
    CARDS = {"VideoObject": "视频缩略图卡", "Article": "文章富卡", "Product": "商品卡",
             "BreadcrumbList": "面包屑富展示", "WebSite": "Sitelinks 搜索框", "Recipe": "食谱卡"}
    shown = 0
    for t in dist:
        if t in CARDS:
            _row(ws, r, [t, CARDS[t], "✓ 可获", "保持必填字段完整"], status_cols={3: "✓ 可获"}); r += 1; shown += 1
    if not shown:
        _row(ws, r, ["—", "暂无可识别的富媒体卡机会", "—", "补充 schema.org 结构化数据"]); r += 1


def _sheet_links(wb, d):
    ws = wb.create_sheet("06 内链结构")
    _widths(ws, [10, 14, 50, 10])
    r = _title_row(ws, "内链结构 / 抓取深度 / 锚文本", 4)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ws.cell(r, 1, "抓取深度分布(从入口点击层数)"); c.font = FONT_LBL
    c.fill = PatternFill("solid", fgColor="E8F1F2"); r += 1
    _header(ws, r, ["深度", "页面数", "说明", ""]); r += 1
    DESC = {0: "入口", 1: "一级(频道/榜单)", 2: "二级(详情/列表)", 3: "三级(章节/播放)", 4: "更深", 5: "最深"}
    dd = d.get("depth_dist", {})
    for depth in sorted(dd.keys(), key=lambda x: int(x)):
        _row(ws, r, [int(depth), dd[depth], DESC.get(int(depth), "更深层级"), ""]); r += 1
    r += 1
    s = d.get("summary", {})
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ws.cell(r, 1, f"孤儿页:{s.get('orphan_pages',0)}  |  最大深度:{s.get('max_depth',0)}  |  "
                      f"内链总数(抽样):{s.get('internal_links',0):,}  |  通用/空锚文本:{s.get('generic_anchor_pct',0)}%")
    c.font = FONT_CELL; c.border = BORDER; c.alignment = LEFT; ws.row_dimensions[r].height = 24


def _sheet_templates(wb, d):
    ws = wb.create_sheet("07 各模板评估")
    _widths(ws, [16, 9, 9, 9, 9, 14, 18, 40])
    r = _title_row(ws, "各模板 SEO 评估", 8)
    _header(ws, r, ["模板", "页面数", "通过项", "警告", "不通过", "sitemap覆盖", "Web Vitals", "主要问题"]); r += 1
    tmpls = d.get("templates", {})
    cwv = d.get("cwv", {})
    for name, info in sorted(tmpls.items(), key=lambda kv: -kv[1]["pages"]):
        top_issues = sorted(info.get("issues", {}).items(), key=lambda kv: -kv[1])[:3]
        issue_str = "、".join(f"{k}×{v}" for k, v in top_issues) or "—"
        cv = cwv.get(name, {}) if isinstance(cwv, dict) else {}
        cwv_str = (f"LCP {cv.get('lcp')}ms {cv.get('verdict','')}" if isinstance(cv, dict) and cv.get("lcp") else "未采")
        _row(ws, r, [name, info["pages"], info["pass"], info["warn"], info["fail"], "—", cwv_str, issue_str]); r += 1


def _sheet_issues(wb, syn):
    ws = wb.create_sheet("08 关键问题清单")
    _widths(ws, [10, 10, 14, 40, 40, 44])
    r = _title_row(ws, "关键问题清单(按优先级)", 6)
    _header(ws, r, ["编号", "优先级", "类别", "问题描述", "影响", "修复建议"]); r += 1
    issues = (syn.get("issues") if isinstance(syn, dict) else None) or []
    if not issues:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(r, 1, "(本次未识别出 P0-P2 关键问题)").font = Font(name=_F, size=10, italic=True, color="6B7280"); r += 1
    for i, it in enumerate(issues, 1):
        if not isinstance(it, dict):
            continue
        pri = it.get("priority") or it.get("优先级") or "P2"
        _row(ws, r, [it.get("issue_id") or f"SEO-{i:02d}", pri, it.get("module") or it.get("类别") or "",
                     it.get("title") or it.get("current_behavior") or it.get("问题描述") or "",
                     it.get("impact_scope") or it.get("影响") or it.get("expected_behavior") or "",
                     it.get("fix_suggestion") or it.get("修复建议") or ""],
             status_cols={2: "不通过" if str(pri) in ("P0",) else ("警告" if str(pri) in ("P1",) else "")})
        ws.row_dimensions[r].height = max(30, len(str(it.get("title", ""))) // 2 + 18); r += 1
    # 优点
    strengths = (syn.get("strengths") if isinstance(syn, dict) else None) or []
    if strengths:
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(r, 1, "已验证的优点"); c.font = FONT_LBL; c.fill = PatternFill("solid", fgColor=C_PASS_BG); r += 1
        for sgood in strengths:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            c = ws.cell(r, 1, f"✓ {sgood}"); c.font = Font(name=_F, size=10, color=C_PASS_FG)
            c.alignment = LEFT; c.border = BORDER; r += 1


def _sheet_pages(wb, d):
    ws = wb.create_sheet("09 全页明细")
    _widths(ws, [6, 48, 14, 8, 9, 9, 6, 6, 10, 10, 9, 14, 10])
    pages = d.get("pages", [])
    r = _title_row(ws, f"全部 {sum(1 for p in pages if p.get('status')==200)} 个 200 OK 页面逐页明细", 13)
    _header(ws, r, ["#", "URL", "模板", "深度", "Title长", "Desc长", "KW", "H1", "层级跳级",
                    "图片Alt%", "JSON-LD", "canonical", "P/W/F"]); r += 1
    idx = 0
    for p in pages:
        if p.get("status") != 200:
            continue
        idx += 1
        from urllib.parse import urlparse
        path = urlparse(p.get("url", "")).path or "/"
        _row(ws, r, [idx, path, p.get("template"), p.get("depth"), p.get("title_len"),
                     p.get("desc_len"), p.get("meta_keywords_count"), p.get("h1_count"),
                     "是" if p.get("heading_skip") else "", p.get("alt_pct"),
                     p.get("jsonld_count"), p.get("canonical_issue"),
                     f"{p.get('p_count',0)}/{p.get('w_count',0)}/{p.get('f_count',0)}"],
             zebra=(idx % 2 == 0)); r += 1
    ws.freeze_panes = ws.cell(r - idx, 1)


def _sheet_criteria(wb):
    ws = wb.create_sheet("10 执行与通过标准")
    _widths(ws, [18, 28, 26, 26])
    r = _title_row(ws, "执行标准 & 通过标准", 4)
    _header(ws, r, ["检查项", "通过", "警告", "不通过"]); r += 1
    crit = [
        ("Title", "10-60字唯一", ">60或重复", "缺失"),
        ("Description", "50-160唯一", "异常或重复", "缺失"),
        ("H1", "每页1个含关键词", "—", "缺失/含i18n占位符/多个"),
        ("标题层级", "H1→H2→H3", "跳级", "混乱"),
        ("图片Alt", "≥80%", "50-80%", "<50%"),
        ("canonical", "自指", "跨频道归一", "跨语言/缺"),
        ("JSON-LD", "含 schema", "—", "详情页缺"),
        ("hreflang", "含 x-default", "无 x-default", "缺"),
        ("sitemap 覆盖", "≥80% 关键模板", "30-80%", "<30% 或文件缺"),
        ("EN 本地化", "title/h1/desc 全英文", "部分英文", "全中文"),
        ("LCP", "≤2500ms", "2500-4000ms", ">4000ms"),
        ("CLS", "≤0.1", "0.1-0.25", ">0.25"),
        ("HSTS/CSP", "均设置", "部分设置", "均缺"),
    ]
    for name, g, w, b in crit:
        _row(ws, r, [name, g, w, b]); r += 1
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ws.cell(r, 1, "执行方法"); c.font = FONT_LBL; c.fill = PatternFill("solid", fgColor="E8F1F2"); r += 1
    _header(ws, r, ["环节", "工具/方法", "", ""]); r += 1
    methods = [
        ("全站抓取+内链图谱", "httpx 异步 BFS,深度+入链+死链"),
        ("静态 SEO 30+ 项", "lxml 解析 title/meta/H 标签/img/canonical/og/twitter/hreflang/JSON-LD"),
        ("结构化数据校验", "按 schema.org 统计 @type,识别 @graph 嵌套"),
        ("内链锚文本质量", "内链词法分析,识别通用词与空锚文本"),
        ("Core Web Vitals", "Playwright + PerformanceObserver(LCP/CLS/FCP/TTFB/Load)+资源"),
        ("SSR 完整性", "Playwright 关 JS 渲染,对比可见文本比例"),
        ("sitemap 验证", "下载分片解析 URL + lastmod + 与爬取集对齐"),
        ("技术 SEO", "响应头检查 HTTP 版本/TLS/压缩/HSTS/CSP/nosniff/cache"),
        ("综合分析", "AI 基于以上真实采集数据生成总览与关键问题清单"),
    ]
    for env, method in methods:
        _row(ws, r, [env, method, "", ""]); r += 1
