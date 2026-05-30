"""弱网/断网测试报告 → Excel(多档位矩阵 + 逐档行为 + 韧性 + 问题清单 + 标准)。

输入:
  report      : LLM 报告(verdict / issues)
  network_data: NetProbeData.to_dict() —— 多网络档位真实探测事实层
  meta        : {site/target, run_id, model, tested_at}
沿用 SEO 报告同款视觉风格(同配色/表头/状态三色)。
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# 复用 seo_excel 的样式 helper,保持两套 Excel 视觉一致
from packages.reporting.seo_excel import (
    _F, _header, _row, _status_fill, _title_row, _widths,
    C_PASS_BG, C_PASS_FG, FONT_CELL, FONT_LBL,
)


def build_network_xlsx(report: dict[str, Any], network_data: dict[str, Any], meta: dict[str, Any]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    _sheet_overview(wb, report, network_data, meta)
    _sheet_matrix(wb, network_data)
    _sheet_behavior(wb, network_data)
    _sheet_issues(wb, report)
    _sheet_criteria(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sheet_overview(wb, report, nd, meta):
    ws = wb.create_sheet("01 综合总览")
    _widths(ws, [16, 24, 24, 24, 24])
    profs = nd.get("profiles", [])
    npass = sum(1 for p in profs if p.get("verdict") == "通过")
    nfail = sum(1 for p in profs if p.get("verdict") == "不通过")
    rec = nd.get("recovery", {}) or {}
    sub = (f"测试时间:{meta.get('tested_at','')}  |  目标:{nd.get('url') or meta.get('site','')}  |  "
           f"档位:{len(profs)}(通过 {npass}/不通过 {nfail})  |  模型:{meta.get('model','')}")
    _title_row(ws, "弱网 / 断网容错测试报告", 5, sub)
    rows = [
        ("综合评定", report.get("verdict_summary") or report.get("verdict") or "(见各档位明细)"),
        ("断网表现", ("断网时给出友好错误提示" if rec.get("offline_has_error_ui")
                      else ("断网时有内容但无明确错误提示" if rec.get("offline_text_len", 0) > 40 else "断网时白屏/无降级"))),
        ("恢复能力", ("网络恢复后自动恢复正常" if rec.get("recovered") else "网络恢复后未自动恢复(需手动刷新或仍异常)")),
        ("后续动作", "；".join(str(i.get("fix_suggestion") or "") for i in (report.get("issues") or [])[:3] if i.get("fix_suggestion")) or "(无)"),
    ]
    r = 4
    for lbl, val in rows:
        lc = ws.cell(r, 1, lbl); lc.font = FONT_LBL; lc.alignment = Alignment(vertical="top", wrap_text=True)
        lc.fill = PatternFill("solid", fgColor="E8F1F2")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        vc = ws.cell(r, 2, str(val)); vc.font = FONT_CELL; vc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(40, min(140, len(str(val)) // 4 + 24))
        r += 1


def _sheet_matrix(wb, nd):
    ws = wb.create_sheet("02 网络档位实测矩阵")
    _widths(ws, [16, 10, 8, 10, 10, 10, 10, 10, 12])
    r = _title_row(ws, "各网络档位真实加载实测", 9)
    _header(ws, r, ["网络档位", "到达", "状态", "加载ms", "FCP ms", "可见字符", "资源数", "控制台错误", "判定"]); r += 1
    for p in nd.get("profiles", []):
        _row(ws, r, [
            p.get("profile"), "是" if p.get("reached") else "否", p.get("status") or "—",
            p.get("load_ms") if p.get("load_ms") is not None else "—",
            p.get("fcp_ms") if p.get("fcp_ms") is not None else "—",
            p.get("text_len", 0), p.get("resources", 0), p.get("console_errors", 0),
            p.get("verdict", "未测"),
        ], status_cols={9: p.get("verdict", "")}); r += 1


def _sheet_behavior(wb, nd):
    ws = wb.create_sheet("03 弱网与断网行为")
    _widths(ws, [16, 12, 12, 14, 50])
    r = _title_row(ws, "逐档位行为细节(加载态 / 错误提示 / 备注)", 5)
    _header(ws, r, ["网络档位", "有加载态", "有错误提示", "超时", "备注/现象"]); r += 1
    for p in nd.get("profiles", []):
        note = p.get("note") or ""
        if p.get("verdict") == "不通过" and not note:
            note = "未成功加载或断网白屏" if p.get("profile", "").startswith("断网") else "加载超时/失败"
        _row(ws, r, [p.get("profile"), "是" if p.get("has_spinner") else "否",
                     "是" if p.get("has_error_ui") else "否", "是" if p.get("timed_out") else "否", note]); r += 1
    # 韧性测试
    r += 1
    rec = nd.get("recovery", {}) or {}
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(r, 1, "断网→恢复 韧性测试"); c.font = FONT_LBL; c.fill = PatternFill("solid", fgColor="E8F1F2"); r += 1
    _header(ws, r, ["环节", "结果", "", "", "说明"]); r += 1
    items = [
        ("在线首次加载", "成功" if rec.get("online_ok") else "失败", ""),
        ("断网重载", "有友好错误提示" if rec.get("offline_has_error_ui") else ("有残留内容" if rec.get("offline_text_len", 0) > 40 else "白屏/失败"), f"可见 {rec.get('offline_text_len','?')} 字符"),
        ("恢复在线重载", "自动恢复正常" if rec.get("recovered") else "未恢复", f"可见 {rec.get('recovered_text_len','?')} 字符"),
    ]
    for env, result, note in items:
        rr = _row(ws, r, [env, result, "", "", note], status_cols={2: ("通过" if result in ("成功", "自动恢复正常", "有友好错误提示") else ("不通过" if result in ("失败", "白屏/失败", "未恢复") else "警告"))}); r += 1


def _sheet_issues(wb, report):
    ws = wb.create_sheet("04 关键问题清单")
    _widths(ws, [10, 10, 16, 40, 40, 44])
    r = _title_row(ws, "关键问题清单(按优先级)", 6)
    _header(ws, r, ["编号", "优先级", "档位/场景", "问题描述", "影响", "修复建议"]); r += 1
    issues = report.get("issues") or []
    if not issues:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(r, 1, "(本次未识别出关键问题)").font = Font(name=_F, size=10, italic=True, color="6B7280"); r += 1
    for i, it in enumerate(issues, 1):
        if not isinstance(it, dict):
            continue
        pri = it.get("priority") or "P2"
        _row(ws, r, [it.get("issue_id") or f"NET-{i:02d}", pri, it.get("module") or "",
                     it.get("title") or it.get("current_behavior") or "",
                     it.get("impact_scope") or it.get("expected_behavior") or "",
                     it.get("fix_suggestion") or ""],
             status_cols={2: "不通过" if str(pri) == "P0" else ("警告" if str(pri) == "P1" else "")})
        ws.row_dimensions[r].height = max(28, len(str(it.get("title", ""))) // 2 + 16); r += 1


def _sheet_criteria(wb):
    ws = wb.create_sheet("05 执行与通过标准")
    _widths(ws, [18, 30, 26, 26])
    r = _title_row(ws, "执行标准 & 通过标准", 4)
    _header(ws, r, ["检查项", "通过", "警告", "不通过"]); r += 1
    crit = [
        ("4G 加载", "≤5s 首屏可用", "5-10s", ">10s 或失败"),
        ("弱网(慢3G)加载", "≤12s 可用且有加载态", "12-25s", ">25s 或白屏"),
        ("2G/极弱", "可加载(可降级)", "超时但有提示", "白屏无提示"),
        ("断网态", "友好错误页 + 重试入口", "有残留内容无提示", "白屏/原始报错"),
        ("加载态反馈", "弱网下有 loading/骨架屏", "部分有", "全程无反馈"),
        ("网络恢复", "自动重连/恢复", "需手动刷新", "不恢复"),
        ("控制台错误", "0 未捕获错误", "少量", "大量报错"),
    ]
    for name, g, w, b in crit:
        _row(ws, r, [name, g, w, b]); r += 1
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ws.cell(r, 1, "执行方法"); c.font = FONT_LBL; c.fill = PatternFill("solid", fgColor="E8F1F2"); r += 1
    _header(ws, r, ["环节", "工具/方法", "", ""]); r += 1
    for env, method in [
        ("多档位限速", "Playwright + CDP Network.emulateNetworkConditions(下行/上行/延迟)"),
        ("断网模拟", "CDP offline=true 真实断网"),
        ("加载指标", "真实加载时长 + FCP + 可见文本量 + 资源数 + 控制台错误"),
        ("行为检测", "检测加载态(spinner/骨架)、错误提示文案、是否白屏"),
        ("韧性测试", "在线→断网重载→恢复在线重载,观察自动恢复"),
        ("综合分析", "AI 基于以上真实多档位数据生成问题清单与结论"),
    ]:
        _row(ws, r, [env, method, "", ""]); r += 1
